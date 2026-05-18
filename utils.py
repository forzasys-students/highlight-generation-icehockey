import os
import subprocess
from collections import defaultdict
from urllib.request import urlopen, Request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# =========================================================
# Small utils
# =========================================================

def run(cmd):
    """
    Run a shell command and raise an exception if it fails.
    """
    subprocess.run(cmd, check=True)


def ensure_dir(p):
    """
    Create a directory if it does not already exist.
    """
    os.makedirs(p, exist_ok=True)


def safe_int(*vals, default=0):
    """
    Try to convert the first non-None / valid value to int.
    If all conversions fail, return the provided default.
    """
    for v in vals:
        try:
            if v is None:
                continue
            return int(v)
        except Exception:
            pass
    return default


def safe_str(x):
    """
    Safely convert a value to string.
    Return an empty string for None.
    """
    return "" if x is None else str(x)


# =========================================================
# XLSX export
# =========================================================

REGION_ORDER = {
    "core": 0,
    "core_after": 1,
    "replay": 2,
}


def ms_to_sec(ms):
    """
    Convert milliseconds to seconds rounded to 3 decimal places.
    """
    return round(ms / 1000.0, 3)


def autosize_ws(ws):
    """
    Auto-adjust worksheet column widths based on content length.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))


def export_segments_xlsx(xlsx_path, goals, raw_paths, meta_by_goal, all_by_goal):
    """
    Export all selected segments and summary statistics to an XLSX file.

    Sheets:
    - segments: detailed segment-level export
    - summary_total: total duration by region/type
    - summary_by_goal: duration by goal/region/type
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "segments"

    ws.append([
        "goal_idx", "clip_file",
        "phase_ord", "game_time",
        "score_after_goal", "scorer",
        "region", "type",
        "start_ms", "end_ms", "dur_ms",
        "start_s", "end_s", "dur_s",
        "logo1_s", "logo2_s", "core_end_s", "core_main_end_s",
        "replay_start_s", "replay_end_s", "hard_cut_end_s",
    ])

    summary = defaultdict(int)
    summary_goal = defaultdict(lambda: defaultdict(int))

    for gi in range(len(raw_paths)):
        ev = goals[gi]
        meta = meta_by_goal[gi]
        per = all_by_goal[gi]

        for (typ, reg), segs in sorted(
            per.items(),
            key=lambda x: (REGION_ORDER.get(x[0][1], 99), x[0][0])
        ):
            for s in sorted(segs, key=lambda z: z["a"]):
                ws.append([
                    gi + 1,
                    os.path.basename(raw_paths[gi]),
                    ev.get("phase_ord"),
                    ev.get("game_time"),
                    safe_str(ev.get("score_after_display") or ev.get("score")),
                    safe_str(ev.get("scorer")),
                    reg, typ,
                    int(s["a"]), int(s["b"]), int(s["dur"]),
                    ms_to_sec(s["a"]), ms_to_sec(s["b"]), ms_to_sec(s["dur"]),
                    None if meta["logo1"] is None else ms_to_sec(meta["logo1"]),
                    None if meta["logo2"] is None else ms_to_sec(meta["logo2"]),
                    ms_to_sec(meta["core_end"]),
                    ms_to_sec(meta["core_main_end"]),
                    None if meta["replay_start"] is None else ms_to_sec(meta["replay_start"]),
                    None if meta["replay_end"] is None else ms_to_sec(meta["replay_end"]),
                    ms_to_sec(meta["hard_cut_end"]),
                ])
                k = (reg, typ)
                summary[k] += s["dur"]
                summary_goal[gi][k] += s["dur"]

    autosize_ws(ws)

    ws2 = wb.create_sheet("summary_total")
    ws2.append(["region", "type", "dur_ms", "dur_s"])
    for (reg, typ), dur in sorted(summary.items(), key=lambda x: (REGION_ORDER.get(x[0][0], 99), x[0][1])):
        ws2.append([reg, typ, int(dur), ms_to_sec(dur)])
    autosize_ws(ws2)

    ws3 = wb.create_sheet("summary_by_goal")
    ws3.append(["goal_idx", "region", "type", "dur_ms", "dur_s"])
    for gi in range(len(raw_paths)):
        for (reg, typ), dur in sorted(summary_goal[gi].items(), key=lambda x: (REGION_ORDER.get(x[0][0], 99), x[0][1])):
            ws3.append([gi + 1, reg, typ, int(dur), ms_to_sec(dur)])
    autosize_ws(ws3)

    ensure_dir(os.path.dirname(xlsx_path) or ".")
    wb.save(xlsx_path)


# =========================================================
# Image helpers (ffmpeg-based)
# =========================================================

def get_drawtext_font_arg():
    return "fontfile='assets/Arial.ttf':"

def escape_drawtext(s: str) -> str:
    """
    Escape text so it is safe to embed inside ffmpeg drawtext filter.
    """
    if s is None:
        return ""
    s = str(s)
    s = s.replace("—", "-").replace("–", "-")
    s = s.replace("\\", r"\\")
    s = s.replace(":", r"\:")
    s = s.replace("'", r"\'")
    s = s.replace("%", r"\%")
    s = s.replace("\n", " ")
    return s


def download_image(url, out_path):
    """
    Download an image from URL to a local file.
    """
    ensure_dir(os.path.dirname(out_path) or ".")
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=20) as r:
        raw = r.read()
    with open(out_path, "wb") as f:
        f.write(raw)



# =========================================================
# Shot boundary utils
# =========================================================

def _get_det_label(d):
    """
    Normalize detection label field name from sportsbd output.
    Different model outputs may use different keys.
    """
    if "predicted_class" in d:
        return d["predicted_class"]
    return (d.get("label") or d.get("class") or d.get("pred") or d.get("prediction") or d.get("boundary"))


def boundary_times_ms(detections):
    """
    Extract all boundary-like timestamps in milliseconds:
    - hard cuts
    - fade-ins
    - logo transitions
    """
    times = []
    for d in detections:
        cls = _get_det_label(d)
        if cls in {"hard", "fadein", "logo"}:
            times.append(int(d["timestamp_ms"]))
    return sorted(set(times))


def logo_times_ms(detections):
    """
    Extract only logo transition timestamps in milliseconds.
    """
    logos = []
    for d in detections:
        if _get_det_label(d) == "logo":
            logos.append(int(d["timestamp_ms"]))
    return sorted(set(logos))


# =========================================================
# Interval helpers
# =========================================================

def segs_sorted(segs):
    """
    Sort segment dicts by start and end time.
    """
    return sorted(segs, key=lambda s: (int(s["a"]), int(s["b"])))


def segs_to_intervals(segs, hard_end):
    """
    Convert segment dicts into clipped [start_ms, end_ms] intervals.
    """
    out = []
    for s in segs_sorted(segs):
        a = max(0, int(s["a"]))
        b = min(int(hard_end), int(s["b"]))
        if b > a:
            out.append([a, b])
    return out


def intervals_len(intervals):
    """
    Compute total duration of a list of intervals.
    """
    return sum(b - a for a, b in intervals)
